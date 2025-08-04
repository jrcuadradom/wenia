import difflib
import os
import re
import boto3
import requests
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
import pdfplumber as pdfplumber


ADDRESS_STREET = ['CRA', 'Carrera', 'Calle', 'Cl', 'Transversal', 'Tv']
ADDRESS_NUM = ['#','No', 'Num', 'Numero']
ADDRESS_DIV = ['-',' ']
wb_name = "addresses.xlsx"
wb_accuracy = "addresses_validated.xlsx"
NORMAL_ADDRESS = {'CRA': 'CRA', 'Kra': 'CRA', 'Carrera': 'CRA', '#': '#','No': '#', 'Num': '#', 'Calle': 'Calle','Cl': 'Calle', 'Tv': 'Transversal', 'Transversal': 'Transversal','Numero': '#'}
map_html= "map.html"
file_path = "Documents/"



def connect_to_bucket():
    """
    Connect to S3 bucket using the keys stored in keys.env

    return: the S3 client object
    """
    load_dotenv('keys.env')
    aws_key = os.getenv("AWS_KEY")
    aws_secret_key = os.getenv("AWS_SECRET_KEY")
    s3_client = boto3.client('s3',
                               aws_access_key_id=aws_key,
                               aws_secret_access_key=aws_secret_key,
                               region_name='us-west-1')

    return s3_client

def upload_file_to_s3(document):
    """
    Push the initial documents to S3 bucket

    :parameter:
    document(str): process the document name from process_pdf()

    :return:
    """
    s3_client = connect_to_bucket()
    bucket = os.getenv("BUCKET")
    try:
        s3_client.upload_file(document, bucket, document)
        print('Success')
    except Exception as e:
        print(e)

def process_pdf():
    """
    Process all the PDF documents stored in Documents folder and call upload_file_to_s3() to push the files.
    Get the address in the document using a regular expression. Call find_homonyms(), similarity_check() and load_map()

    :return:
    """
    for file in os.listdir(file_path):
        if file.lower().endswith(".pdf"):
            path = os.path.join(file_path, file)
            upload_file_to_s3(path)
            with pdfplumber.open(path) as pdf:
                page = pdf.pages[0]
                text = page.extract_text()
                if text:
                    documentaddress = re.search(r"(?i)Dirección:\s*(.*)", text)
                    if documentaddress:
                        address = documentaddress.group(1).strip()
                        filename= format_filename(address)
                        find_homonyms(address)
                        similarity_check(address,filename)
                        load_map(filename)
                        print(address)
                    else:
                        print('No address found')



def format_filename(address):
    """
    Format the address to have a suffix to create files for each address.
    :param address: Takes the address got from the PDF file
    :return: A string with no special characters nor spaces
    """
    filename = address.replace('#', '').replace('-', '').replace(' ','')
    filename = filename.replace(' ', '_')
    return filename


def find_homonyms(content):
    """
    Find all the homonyms using the list of possible values.
    :param content: The address got from the PDF document
    :return: A list of all homonyms from the original address
    """
    homonyms = []
    split_address = content.replace('#', '').replace('-','').split()
    for x in ADDRESS_STREET:
        for y in ADDRESS_NUM:
            for z in ADDRESS_DIV:
                homonyms_address = x + ' ' + split_address[1] + ' '+ y + ' '+ split_address[2] + ' ' + z + ' '+ split_address[3]
                homonyms.append(homonyms_address.replace('   ',' '))
    filename = format_filename(content)
    homonyms_to_excel(homonyms, 'homonyms', filename)
    return homonyms

def homonyms_to_excel(data, type,address):
    """
    Create excel files for homonyms and accuracy using a list
    :param data: list of addresses either after the homonyms process or the accuracy validation
    :param type: "homonyms" or "accuracy"
    :param address: The address without special characters nor white spaces to set the name of the file
    :return:
    """
    wb = Workbook()
    ws = wb.active
    index = 1
    #while index <= len(data):
    if(type == 'homonyms'):
        for j in data:
            ws['A'+str(index)] = j
            index+=1
        wb.save("Documents/" + address + wb_name)
        upload_file_to_s3("Documents/" + address + wb_name)
    elif(type == 'accuracy'):
        for j in data:
            ws['A' + str(index)] = j[0]
            ws['B' + str(index)] = j[1]
            ws['C' + str(index)] = j[2]
            index += 1
        wb.save("Documents/" + address + wb_accuracy)
        upload_file_to_s3("Documents/" + address + wb_accuracy)


def normalize_address(address):
    """
    Normalize the address using the NORMAL_ADDRESS dictionary as the address are the same
    :param address: the original address
    :return: the normalized address i.e. KRA -> Carrera, CRA -> Carrera
    """
    split_address = address.replace('-', '').split()
    normal_address = NORMAL_ADDRESS[split_address[0]] + ' ' + split_address[1]+ ' ' + NORMAL_ADDRESS[split_address[2]] +' ' + split_address[3] + ' - ' + split_address[4]
    return normal_address

def similarity_check(address_original,filename):
    """
    Check the similarity of original address against the homonyms both normalized to increase to calculate the similarity using difflib. It sends an array of arrays to homonyms_to_excel
    :param address_original: The address of the document
    :param filename: The formatted name for the excel file
    :return:
    """
    accuracy_list=[]
    all = []
    address_original_low = normalize_address(address_original).lower()
    wb = load_workbook("Documents/"+filename+wb_name)
    sheet = wb.active
    for row in range(1,sheet.max_row+1):
        value = sheet['A'+str(row)].value
        normal_address = normalize_address(value).lower()
        similitud = difflib.SequenceMatcher(None,address_original_low, normal_address).ratio()
        #to find the similarity without normalized address
        #similitud = difflib.SequenceMatcher(None, address_original_low, value.lower()).ratio()
        if(similitud > 0.9):
            accuracy_list.append(value)
            accuracy_list.append(similitud)
            coords = get_geo_location(value)
            accuracy_list.append(str(coords[0]) +', ' + str(coords[1]))
            all.append(accuracy_list.copy())
            accuracy_list.clear()
    homonyms_to_excel(all, 'accuracy', filename)
    #return accuracy_list

def remove_local_files():
    """
    Remove all the local files created as the files are pushed to S3(Not activated to check the results without S3 if needed)
    :return:
    """
    for file in os.listdir(file_path):
        if file.endswith('.xlsx') or file.endswith('.html'):
            full_path = os.path.join(file_path,file)
            try:
                os.remove(full_path)
            except Exception as e:
                print(e)

def get_geo_location(address):
    """
    Get the coordinates of an address using the Geocoding API of google
    :param direccion: the address to be calculated in the API
    :return:the lat and lng of the address
    """
    load_dotenv('keys.env')
    api_key = os.getenv("API_KEY")
    url = "https://maps.googleapis.com/maps/api/geocode/json"
    params = {
            "address": address+', Bogota',
            "key": api_key
    }
    respond = requests.get(url, params=params)
    data = respond.json()
    coord = data['results'][0]['geometry']['location']
    return coord['lat'], coord['lng']


def load_map(address):
    """
    Create a html file with the coordinates
    :param address: The address to be displayed in the map
    :return:
    """
    load_dotenv('keys.env')
    map_key = os.getenv("MAP_KEY")
    checked_coors = []
    wb = load_workbook("Documents/"+address+wb_accuracy)
    sheet = wb.active
    for row in range(1, sheet.max_row + 1):
        value = sheet['C' + str(row)].value
        if value not in checked_coors:
            checked_coors.append(value)
    with open(address+map_html,'w') as f:
        f.write(f"""
        <html>
  <head>
    <title>Add a Map with Markers using HTML</title>

    <link rel="stylesheet" type="text/css" href="./style.css" />
    <script type="module" src="./index.js"></script>
  </head>
  <body>
    <gmp-map
      center="4.6548449,-74.1586209"
      zoom="10"
      map-id="DEMO_MAP_ID"
      style="height: 500px"
    >""")
        for coor in checked_coors:
            f.write(f"""<gmp-advanced-marker
            position="{coor}"
            ></gmp-advanced-marker>
            """)
        f.write(f"""</gmp-map>
    <script
      src="https://maps.googleapis.com/maps/api/js?key={map_key}&libraries=maps,marker&v=beta"
      defer
    ></script>
  </body>
</html>
        """)

if __name__ == '__main__':
    process_pdf()





